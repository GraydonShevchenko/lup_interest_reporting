# Filename: lup_interest_reporting.py
# Author: Graydon Shevchenko
# Created: July 17, 2025
# Description: This script is designed to report out on overlas of features with a given area of interest to help support landuse planning

import sys
import os
import logging
import arcpy
import pandas as pd
import warnings
import openpyxl
import re
import traceback

from copy import copy
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from argparse import ArgumentParser
from collections import defaultdict
from datetime import datetime


script_dir = os.path.dirname(os.path.abspath(__file__))

from util.environment import Environment


def run_app() -> None:
    """
    FUNCTION

    run_app: Used to control the general flow of the script 
    """

    # Gather the input parameters and set up the class oject
    file_num, out_dir, xls, aoi, aoi_fld, leave, logger = get_input_parameters()
    lup_overlaps = LUP_Overlaps(file_number=file_num, output_dir=out_dir, xls_schema=xls, aoi=aoi, aoi_field=aoi_fld,
                            leave_areas=leave, logger=logger)
    
    # Run the class methods needed to perform the overlap assessment
    lup_overlaps.setup_aoi()
    lup_overlaps.overlay_values()
    lup_overlaps.write_excel()

    # Delete the class object
    del lup_overlaps


def get_input_parameters() -> tuple:
    """
    FUNCTION

    get_input_parameters: Uses the argparse library to gather the input parameters and set up the logging object

    Returns:
        tuple: file name/number, output directory, excel schema, area of interest feature layer, area of interest field, leave areas feature layer, directory the script is run from, logger object
    """

    try:
        # Set up the input parameters. This matches what is put into the toolbox tool within ArcGIS.  Doing this allows for running the script within command line
        parser = ArgumentParser(description='This script is used to calculate land use planning overlap values '
                                            'based on an input area of interest. Regional based CE values can be used ' 'by inputting a schema using an excel file')
        parser.add_argument('file_num', type=str, help='File Name or number')
        parser.add_argument('out_dir', type=str, help='Output directory')
        parser.add_argument('xls', type=str, help='Input schema Excel file')
        parser.add_argument('aoi', type=str, help='Area of interest layer')
        parser.add_argument('aoi_fld', type=str, help='AOI ID field')
        parser.add_argument('leave', type=str, help='AOI leave areas')
        parser.add_argument('--log_level', default='INFO', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR'], 
                            help='Log level for message output')
        parser.add_argument('--log_dir', help='Path to the log file directory')

        args = parser.parse_args()
       

        # Set up the logger object
        logger = Environment.setup_logger(args)

        return args.file_num, args.out_dir, args.xls, args.aoi, args.aoi_fld, args.leave, logger

    except Exception as e:
        logging.exception(f'Unexpected exception, program terminating: {repr(e)}')


class LUP_Overlaps:
    """
    CLASS

    Main class object that contains the methods required to complete the cumulative effects framework analysis
    """
    def __init__(self, file_number:str, output_dir:str, xls_schema:str, aoi:str, aoi_field:str, leave_areas:str, 
                 logger:logging.Logger) -> None:
        """
        CLASS METHOD

        __init__: Initialization function that is run when this class object is created.  Its sets up the class variables required as well as runs initial processes to set up the workspace and read through the excel schema
        Args:
            file_number (str): the file name or number which is used to help name the output file and geodatabase
            output_dir (str): directory where the script outputs will be stored
            xls_schema (str): path to the excel schema document
            aoi (str): area of interest feature layer
            aoi_field (str): area of interest field
            leave_areas (str): leave areas feature layer
            python_dir (str): directory where the script is run from
            logger (logging.Logger): logger object for messaging
        """

        # Write the parameters into the class variables
        self.file_number = file_number.replace(' ', '_')
        self.output_dir = output_dir
        self.xls_schema = xls_schema
        self.aoi = aoi
        self.fld_aoi = aoi_field if aoi_field != '#' else None
        self.leave_areas = leave_areas if leave_areas != '#' else None
        self.logger = logger

        # Set up the lup dictionary for storing the values
        self.dict_lup_values = defaultdict(lambda: defaultdict(lambda: LU_Value))

        # File paths for spatial files, connections and output files
        # self.logger.info('Creating bcgw connection')
        # self.bcgw_db = oracledb.connect(dsn='bcgw.bcgov/idwprod1.bcgov', user=self.bcgw_un, password=self.bcgw_pw)
        self.bcgw = os.path.join(os.path.dirname(sys.argv[0]), 'util', 'BCGW_Connection.sde')
        # self.bcgw = Environment.create_bcgw_connection(location=self.output_dir, bcgw_user_name=self.bcgw_un,
        #                                                bcgw_password=self.bcgw_pw, logger=self.logger)

        # try:
        #     self.bcgw_db.ping()
        #     self.logger.info('Connection is active and database is reachable.')
        # except oracledb.Error as e:
        #     self.logger.error(f'Connection is not active or database is unreachable: {e}')
        #     sys.exit(1)
        
        self.out_gdb = os.path.join(self.output_dir, f'lup_data_{self.file_number}.gdb')
        self.temp_gdb = 'memory'
        self.fd_incoming = os.path.join(self.out_gdb, 'incoming')
        self.fd_work = os.path.join(self.out_gdb, 'work')

        self.fc_aoi = os.path.join(self.fd_incoming, 'aoi')
        self.fc_leave_areas = os.path.join(self.fd_incoming, 'leave_areas')
        self.fc_net_aoi = os.path.join(self.fd_incoming, 'net_aoi')



        self.aoi_total = 0
        self.dict_aoi_area = defaultdict(int)

        self.xls_output = os.path.join(self.output_dir, 
                                      f'LUP_Overview_{self.file_number}_{datetime.strftime(datetime.today(), "%Y%m%d")}.xlsx')

        # Other variables required for running the script
        self.xl_style = None
        self.str_summary = 'Summary'
        self.str_overall = 'AOI Overall'

        self.fld_area = 'AreaHA'

        # Standard headers to inlcude in the excel output for every lup value
        self.standard_poly_headers = ['Assessment Unit', 'Area (ha) of Assessment Unit', 
                                 'Area (ha) of Assessment Unit overlap with AOI', 
                                 '% of Assessment Unit that Overlaps with AOI', 
                                 '% of AOI that Overlaps with Assessment Unit']
        self.standard_line_headers = ['Assessment Unit', 'Length (m) of Assessment Unit', 
                                 'Length (m) of Assessment Unit overlap with AOI', 
                                 '% of Assessment Unit that Overlaps with AOI', 
                                 '% of AOI that Overlaps with Assessment Unit']
        self.standard_point_headers = ['Assessment Unit', 'Total Count of Assessment Unit', 'Count within AOI', 
                                 '% of Assessment Unit that Overlaps with AOI', 
                                 '% of AOI that Overlaps with Assessment Unit']
        self.summary_headers = ['Number of Assessment Units', 
                                'Total Area (ha), Length (m) or Count of Features', 'Area (ha), Length (m) or Count of Assessment Unit with AOI',
                                '% of Assessment Unit that Overlaps with AOI', 
                                '% of AOI that Overlaps with Assessment Unit']

        arcpy.env.overwriteOutput = True
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

        # Set up the output geodatabase in the specified output directory
        try:
            self.logger.info(f'Setting up geodatabase {os.path.basename(self.out_gdb)}')
            arcpy.management.CreateFileGDB(out_folder_path=os.path.dirname(self.out_gdb), 
                                           out_name=os.path.basename(self.out_gdb))
        except:
            pass

        # Create the incoming and working feature datasets within the new geodatabase
        self.logger.info('Adding in feature datasets')
        if arcpy.Exists(self.fd_incoming):
            arcpy.management.Delete(in_data=self.fd_incoming)
        try:
            arcpy.management.CreateFeatureDataset(out_dataset_path=os.path.dirname(self.fd_incoming), 
                                                  out_name=os.path.basename(self.fd_incoming), spatial_reference=3005)
        except:
            pass
        if arcpy.Exists(self.fd_work):
            arcpy.management.Delete(in_data=self.fd_work)
        try:
            arcpy.management.CreateFeatureDataset(out_dataset_path=os.path.dirname(self.fd_work), 
                                                  out_name=os.path.basename(self.fd_work), spatial_reference=3005)
        except:
            pass

        # Copy the area of interest and leave area feature layers to the geodatabase if they exist
        self.logger.info('Copying data to working geodatabase')
        if self.aoi:
            self.logger.info(f'Copying aoi: {os.path.basename(self.aoi)}')
            arcpy.management.CopyFeatures(in_features=self.aoi, out_feature_class=self.fc_aoi)

        if self.leave_areas:
            self.logger.info(f'Copying leave areas: {os.path.basename(self.leave_areas)}')
            arcpy.management.CopyFeatures(in_features=self.leave_areas, out_feature_class=self.fc_leave_areas)
        
        # Set the processing extent to the area of interest, helps with exporting only needed features instead of the whole input dataset
        

        # Read in the schema values from the input excel file
        self.logger.info('Reading in lup values from excel schema')
        try:
            lup_df = pd.read_excel(self.xls_schema, sheet_name='LUP Indicators', engine='openpyxl', skiprows=[1], 
                                    na_filter=False)
        except:
            self.logger.error('Could not read in the excel schema as it does not contain the LUP Indicators sheet')
            sys.exit()

        try:
            # Loop through each row and gather the lup value information
            for row in lup_df.itertuples():
                # Check if any of the required fields are empty, if so, then skip over it
                if any([not i for i in [row.CATEGORY, row.DATASET_NAME, row.PATH]]):
                    self.logger.warning(f'The value {row.DATASET_NAME} is missing one or more of the required fields; skipping this value')
                    continue
                # Check the indicated path object exists, if not it may be a BCGW path
                if arcpy.Exists(row.PATH):
                    full_path = row.PATH
                else:
                    # Check if the BGW path exists, if not, then skip using the value in the report
                    if arcpy.Exists(os.path.join(self.bcgw, row.PATH)):
                        full_path = os.path.join(self.bcgw, row.PATH)
                    else:
                        self.logger.warning(f'!!! Could not find the path specified in the excel: {row.PATH}. Skipping this     value')
                        continue

                lst_fields = [f.name for f in arcpy.ListFields(dataset=full_path)]
                id_fields = str(row.UNIQUE_ID_FIELD).replace(' ','').split(',') if row.UNIQUE_ID_FIELD else []
                assess_fields = str(row.ASSESSMENT_UNIT_FIELD).replace(' ','').split(',') \
                    if row.ASSESSMENT_UNIT_FIELD else []
                
                fc_fields = id_fields + assess_fields
                drop_fields = []
                for fld in fc_fields:
                    if fld not in lst_fields:
                        drop_fields.append(fld)

                drop_fields = list(set(drop_fields))
                if drop_fields:
                    self.logger.warning(f'The following specifed fields were not found in the {row.DATASET_NAME} dataset: {drop_fields}. They will be removed from the analysis which may result in unexpected summarizations.')
                    id_fields = [f for f in id_fields if f not in drop_fields]
                    assess_fields = [f for f in assess_fields if f not in drop_fields]


                # Check the indicated path object exists, if not it may be a BCGW path
                join_path_type = None
                join_path = ''
                if row.JOIN_TABLE_PATH != '':
                    if arcpy.exists(row.JOIN_TABLE_PATH):
                        join_path = row.JOIN_TABLE_PATH
                    else:
                        # Check if the BCGW path exists, if not, then skip using the value in the report
                        if arcpy.exists(os.path.join(self.bcgw, row.JOIN_TABLE_PATH)):
                            join_path = os.path.join(self.bcgw, row.JOIN_TABLE_PATH)
                        else:
                            self.logger.warning(f'!!! Could not find data at the path specified in the excel: {row. JOIN_TABLE_PATH}. Skipping this value')
                            continue


                # Create a CE Value object with the gathered information and add it to the dictionary
                lup_value = LU_Value(name=row.DATASET_NAME, category=row.CATEGORY, 
                                    path=full_path, id_fields=id_fields, 
                                    assess_fields=assess_fields, sql=row.SQL, source_field=row.SOURCE_FIELD, join_table_path=join_path, join_table_type=join_path_type, join_table_field=row.JOIN_TABLE_FIELD, buffer=row.BUFFER)
                self.dict_lup_values[row.CATEGORY][row.DATASET_NAME] = lup_value
        except Exception as e:
            # Exit out of the script if there is an error; most likely caused by an incorrect schema file used
            self.logger.error(f'There was an issue with the input schema dataset used.  Ensure all the required fields exist within the excel sheet.  Stack Trace: {repr(e)}')
            sys.exit()

        # Read in the values from the additional field schema page. This includes formatting and values used for each indicatory on each specified dataset
        try:
            self.logger.info('Reading in additional field schemas')
            wb = openpyxl.load_workbook(filename=self.xls_schema)
            ws = wb['Additional Fields']
        except:
            self.logger.error('Could not read in the excel schema as it does not contain the Additional Fields sheet')
            sys.exit()

        # Loop through each row of the excel sheet
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            # Pull in the row information into a Field Schema object
            ds_name = row[0].value
            fld_schema = FieldSchema()
            fld_schema.name = row[1].value
            fld_schema.label = row[2].value if row[2].value else row[1].value
            fld_schema.other_fields = row[3].value.replace(' ','').split(',') if row[3].value else []
            fld_schema.value_type = row[4].value

            # Loop through the columns containing the break values and formatting for each lup value
            for i in range(5, len(row)):
                if not row[i].value:
                    break
                # Set up a Value Schema object
                cell_schema = ValueSchema()
                cell_value = row[i].value

                # If the value type specified is a range, then parse the values and place them in the value schema object
                if fld_schema.value_type == 'Range':
                    # Pull out the numerical values from the string and place them in a list.  This is used to extract ranges that are indicated using the format low value-high value (eg. 0.5-1.2)
                    lst_range = re.findall(r'-?\d+\.?\d*--?\d+\.?\d*', cell_value.replace(' ',''))
                    range_operator = None

                    # If the extracted values are in fact a range, then place them in applicable variables
                    if lst_range:
                        for rng in lst_range:
                            low_val = rng.split('-')[0]
                            high_val = rng.split('-')[1]
                            break
                    # If it is not a standard range, then search for the comparison operators and parse out appropriately
                    else:
                        range_val = cell_value.replace(' ','')
                        if '<=' in range_val:
                            low_val = None
                            high_val = range_val.replace('<=','')
                            range_operator = '<='
                        elif '>=' in range_val:
                            low_val = range_val.replace('>=','')
                            high_val = None
                            range_operator = '>='
                        elif '<' in range_val:
                            low_val = None
                            high_val = range_val.replace('<','')
                            range_operator = '<'
                        elif '>' in range_val:
                            low_val = range_val.replace('>','')
                            high_val = None
                            range_operator = '>'

                    # Account for percent values
                    low_val = None if not low_val else float(low_val) if '%' not in low_val else float(low_val.replace('%',''))/100
                    high_val = None if not high_val else float(high_val) if '%' not in high_val else float(high_val.replace('%',''))/100

                    # Add to the cell schema object
                    cell_schema.range_low = low_val
                    cell_schema.range_high = high_val
                    cell_schema.range_operator = range_operator

                # Pull out the cell formatting and place in the cell schema object
                cell_schema.style_align = copy(row[i].alignment)
                cell_schema.style_border = copy(row[i].border)
                cell_schema.style_fill = copy(row[i].fill)
                cell_schema.style_font = copy(row[i].font)
                cell_schema.style_format = copy(row[i].number_format)

                # Add the cell schema object to the field schema object
                fld_schema.dict_values[cell_value] = cell_schema

            # Find the appropriate dataset in the lup values dictionary and add in the field schema object
            for lup in self.dict_lup_values:
                for ds in self.dict_lup_values[lup]:
                    if ds ==  ds_name:
                        self.dict_lup_values[lup][ds].other_fields_schema[fld_schema.name] = fld_schema
        wb.close()

    def __del__(self):

        # Environment.delete_bcgw_connection(location=self.output_dir, logger=self.logger)
        self.logger.info('Clearing workspace cache')
        arcpy.management.ClearWorkspaceCache()

        

    def setup_aoi(self) -> None:
        """
        CLASS METHOD

        setup_aoi: Sets up the area of interest datset and removes any leave areas if required
        """
        # Add in the total area field and calcualte the area in hectares for each feature
        self.logger.info('Determining aoi area in hectares')
        arcpy.management.AddField(in_table=self.fc_aoi, field_name=self.fld_area, field_type='DOUBLE')
        with arcpy.da.UpdateCursor(in_table=self.fc_aoi, field_names=[self.fld_area, 'SHAPE@AREA']) as u_cursor:
            for row in u_cursor:
                row[0] = row[1]/10000
                u_cursor.updateRow(row)

        # If the leave areas exist, then erase them out of the area of interest.  Uses the union tool as the Erase requires an Advanced licence
        if arcpy.Exists(self.fc_leave_areas):

            self.logger.info('Removing leave areas')
            # Erase leave areas from the aoi
            arcpy.analysis.PairwiseErase(in_features=self.fc_aoi, erase_features=self.fc_leave_areas, 
                                         out_feature_class=self.fc_net_aoi)
        else:
            # Copy the aoi if no leave areas were specified
            arcpy.management.CopyFeatures(in_features=self.fc_aoi, out_feature_class=self.fc_net_aoi)

        # Loop through the net aoi dataset and add the total areas to a dictionary for later use
        lst_fields = ['SHAPE@AREA']
        if self.fld_aoi:
            lst_fields.append(self.fld_aoi)
        with arcpy.da.SearchCursor(in_table=self.fc_net_aoi, field_names=lst_fields) as s_cursor:
            for row in s_cursor:
                self.aoi_total += row[0] / 10000
                self.dict_aoi_area[self.str_overall] += row[0] / 10000
                if self.fld_aoi:
                    self.dict_aoi_area[row[1]] += row[0] / 10000


    # def fetch_oracle_geodata(self, connection, table, id_flds, aoi_gdf, sql_filter=None, logger=None):
    #     """
    #     Fetch spatial data from an Oracle database, returning a GeoDataFrame.
    #     Attempts WKB first for speed + curved geometry handling, falls back to WKT if needed.
    #     Automatically tries GEOMETRY, then SHAPE columns.
    #     """

    #     # Dissolve AOI into single geometry
    #     dis_poly = aoi_gdf.union_all()
    #     wkt_polygon = dis_poly.wkt

    #     def build_query(column, use_wkb=True):
    #         if use_wkb:
    #             geom_select = f"SDO_UTIL.TO_WKBGEOMETRY({column}) AS WKB_GEOM"
    #         else:
    #             geom_select = f"SDO_UTIL.TO_WKTGEOMETRY(SDO_UTIL.RECTIFY_GEOMETRY({column}, 0.0001)) AS WKT_GEOM"

    #         base_query = f"""
    #             SELECT {id_flds}, {geom_select}
    #             FROM {table}
    #             WHERE SDO_FILTER({column}, SDO_GEOMETRY(:geom_wkt, 3005), 'querytype=WINDOW') = 'TRUE'
    #               AND SDO_RELATE({column}, SDO_GEOMETRY(:geom_wkt, 3005), 'mask=ANYINTERACT') = 'TRUE'
    #         """
    #         if sql_filter:
    #             base_query += f" AND {sql_filter}"
    #         return base_query

    #     rows, columns = [], []

    #     # Try WKB first on GEOMETRY column
    #     for column in ["GEOMETRY", "SHAPE"]:
    #         try:
    #             query = build_query(column, use_wkb=True)
    #             self.logger.info(f'    - Trying WKB extraction on "{column}" column')
    #             cur = connection.cursor()
    #             cur.execute(query, geom_wkt=wkt_polygon)
    #             columns = [col[0] for col in cur.description]
    #             for r in cur:
    #                 geom_bytes = r[-1].read() if r[-1] is not None else None
    #                 rows.append(r[:-1] + (geom_bytes,))
    #             if rows:
    #                 break
    #         except oracledb.DatabaseError as e:
    #             self.logger.warning(f'    - WKB extraction failed on "{column}" ({e}). Trying WKT fallback.')

    #             # WKT fallback if WKB fails
    #             try:
    #                 query = build_query(column, use_wkb=False)
    #                 cur = connection.cursor()
    #                 cur.execute(query, geom_wkt=wkt_polygon)
    #                 columns = [col[0] for col in cur.description]
    #                 for r in cur:
    #                     wkt_geom = r[-1]
    #                     rows.append(r[:-1] + (wkt_geom,))
    #                 if rows:
    #                     break
    #             except oracledb.DatabaseError as e2:
    #                 self.logger.error(f'    - WKT fallback failed on "{column}" ({e2})')
    #                 continue

    #     if not rows:
    #         self.logger.warning('***No features overlap the area of interest***')
    #         return gpd.GeoDataFrame(columns=columns)

    #     # Load into DataFrame
    #     df = pd.DataFrame(rows, columns=columns)

    #     # Detect if we used WKB or WKT
    #     if "WKB_GEOM" in df.columns:
    #         df["GEOMETRY"] = df["WKB_GEOM"].apply(lambda g: wkb.loads(g) if g else None)
    #         df["GEOMETRY"] = gpd.GeoSeries.from_wkb(df['WKB_GEOM'])
    #         df = df.drop(["WKB_GEOM",], axis=1)
    #     elif "WKT_GEOM" in df.columns:
    #         df["GEOMETRY"] = df["WKT_GEOM"].apply(lambda g: wkt.loads(g) if g else None)
    #         df = df.drop(["WKT_GEOM"], axis=1)

    #     gdf = gpd.GeoDataFrame(df, geometry=df["GEOMETRY"], crs=3005)
    #     gdf = gdf.dropna(subset=["GEOMETRY"])

    #     return gdf

    # def process_and_export_geodata(self, gdf_lup, aoi_gdf, out_gdb, layer_name, buffer=None, logger=None):
    #     """
    #     Cleans, buffers, unions, and exports a GeoDataFrame into a FileGDB layer.
    #     Returns the final GeoDataFrame.
    #     """
    #     if gdf_lup is None or gdf_lup.empty:
    #         self.logger.warning("***Value does not overlap the area of interest***")
    #         return gdf_lup


    #     # Optional buffer
    #     if buffer:
    #         self.logger.info(f" - buffering by {buffer} metres")
    #         gdf_lup = gdf_lup.set_geometry(gdf_lup.geometry.buffer(buffer))


    #     self.logger.info(" - running union")


    #     gdf_union = gpd.overlay(df1=gdf_lup, df2=aoi_gdf, how="union", keep_geom_type=True)
    #     gdf_union = gdf_union.explode(index_parts=False).reset_index(drop=True)


    #     # Clean fields
    #     gdf_union.drop(columns=["FID", "OID", "OBJECTID", "Shape_Area", "Shape_Length"], inplace=True, errors="ignore")


    #     # Metrics (hectares + length in layer units)
    #     gdf_union["Shape_Area"] = gdf_union.geometry.area / 10000.0
    #     gdf_union["Shape_Length"] = gdf_union.geometry.length


    #     # Export
    #     gdf_union.to_file(filename=out_gdb, layer=layer_name, driver="OpenFileGDB")
    #     self.logger.info(f" - completed export: {layer_name}")


    #     return gdf_union




    def overlay_values(self) -> None:
        """
        CLASS METHOD

        overlay_values: Main function that performs the overlays between the aoi and the lup values, then stores the results in the dictionary
        """
        # arcpy.env.extent = self.fc_net_aoi

        fld_id_aoi = f'FID_net_aoi'
        # self.logger.info(self.gdf_net_aoi.columns)
        aoi_geom = None
        with arcpy.da.SearchCursor(self.fc_net_aoi, ['SHAPE@']) as s_cursor:
            for row in s_cursor:
                if aoi_geom:
                    aoi_geom = aoi_geom.union(row[0])
                else:
                    aoi_geom = row[0]
            # fs = FeatureSet(aoi_geom)

        # aoi_geometry = fs.geometries[0]

        # aoi_sedf = pd.DataFrame.spatial.from_featureclass(self.fc_net_aoi)
        # aoi_geom = None
        # # Get a list of all geometry objects from the GeoDataFrame
        # geometry_list = aoi_gdf['SHAPE'].tolist()

        # # Start with the first geometry as the base
        # unioned_geometry = geometry_list[0]

        # # If there are more geometries, loop through the rest and union them
        # if len(geometry_list) > 1:
        #     for geom in geometry_list[1:]:
        #         unioned_geometry = unioned_geometry.union(geom)

        # # aoi_geom = unioned_geometry
        # aoi_geom = Geometry(unioned_geometry.JSON)
        aoi_lyr = arcpy.management.MakeFeatureLayer(in_features=self.fc_net_aoi, out_layer='aoi_lyr')
        # Loop through the lup values extracted from the input excel schema file
        for lup in self.dict_lup_values:
            # Loop through the datasets for each lup value
            for ds in self.dict_lup_values[lup]:

                self.logger.info(f'Running overlay on {ds}')
                lup_ds = self.dict_lup_values[lup][ds]
                in_fc = lup_ds.path
                join_fc = lup_ds.join_path
                buffer = lup_ds.buffer
                sql = lup_ds.sql
                id_flds = ','.join(lup_ds.id_fields)
                if not lup_ds.id_fields:
                    id_flds = 'OBJECTID'

                temp_fc = os.path.join(self.temp_gdb, 'temp_fc')
                intersect_fc = os.path.join(self.temp_gdb, 'intersect_fc')
                buffer_fc = os.path.join(self.temp_gdb, 'buffer_fc')
                erase_fc = os.path.join(self.temp_gdb, 'erase_fc')
                union_name = f"union_{str(lup_ds.name).lower().replace(' ', '_')}"
                union_fc = os.path.join(self.fd_work, union_name)

                
                fc_lyr = arcpy.management.MakeFeatureLayer(in_features=in_fc, out_layer='fc_lyr', where_clause=sql)

                result = int(arcpy.management.GetCount(fc_lyr).getOutput(0))
                arcpy.management.SelectLayerByLocation(in_layer=fc_lyr, overlap_type='INTERSECT',select_features=aoi_lyr)
                result = int(arcpy.management.GetCount(fc_lyr).getOutput(0))

                arcpy.management.CopyFeatures(in_features=fc_lyr, out_feature_class=temp_fc)
                arcpy.management.Delete(in_data=fc_lyr)
                # arcpy.analysis.Select(in_features=in_fc, out_feature_class=temp_fc, where_clause=sql)
                # arcpy.conversion.ExportFeatures(in_features=in_fc, out_features=temp_fc, field_mapping=field_mappings,
                #                                  where_clause=sql)

                result = int(arcpy.management.GetCount(temp_fc).getOutput(0))

                if result == 0:
                    # arcpy.analysis.Select(in_features=in_fc, out_feature_class=os.path.join(self.fd_work, 'temp_fc'), where_clause=sql)
                    self.logger.warning('***Dataset does not overlap AOI***')
                    continue

                fld_id_lu = f"FID_{str(lup_ds.name).lower().replace(' ', '_')}"
                fld_id_lu = fld_id_lu[:60] if len(fld_id_lu) > 60 else fld_id_lu
                arcpy.management.AddField(in_table=temp_fc, field_name=fld_id_lu, field_type='LONG')

                lup_id = 1
                # Use a search cursor with a spatial query to get filtered data
                with arcpy.da.UpdateCursor(temp_fc, ['SHAPE@', fld_id_lu]) as cursor:
                    for row in cursor:
                        if aoi_geom.disjoint(row[0]):
                            cursor.deleteRow()
                        else:
                            row[1] = lup_id
                            lup_id += 1
                            cursor.updateRow(row)

                if buffer:
                    self.logger.info(f'    - buffering by {buffer} metres')
                    arcpy.analysis.PairwiseBuffer(in_features=temp_fc, out_feature_class=buffer_fc,
                                                   buffer_distance_or_field=buffer)
                    arcpy.management.Delete(temp_fc)
                    temp_fc = buffer_fc

                self.logger.info('    - intersecting with aoi')
                arcpy.analysis.PairwiseIntersect(in_features=[self.fc_net_aoi, temp_fc], out_feature_class=intersect_fc)
                self.logger.info('    - erasing aoi')
                arcpy.analysis.PairwiseErase(in_features=temp_fc, erase_features=self.fc_net_aoi,
                                              out_feature_class=erase_fc)
                self.logger.info('    - combining datasets')
                arcpy.management.Merge(inputs=[intersect_fc, erase_fc], output=union_fc, 
                                       field_match_mode='USE_FIRST_SCHEMA')
                
                for fc in [temp_fc, intersect_fc, erase_fc]:
                    arcpy.management.Delete(in_data=fc)

                ds_type = arcpy.Describe(union_fc).shapeType
                self.dict_lup_values[lup][ds].data_type = ds_type
                    

                # If there is a join tables specifed in the schema for this value then, join it to the unioned dataset
                if join_fc:
                    # If the join fields were not specified correctly in the excel schema, then don't perform the join
                    if not any([lup_ds.source_field, lup_ds.join_field]):
                        self.logger.warning('The fields required for the table join were not specified in the schema')
                    else:
                        self.logger.info(f'Joining {os.path.basename(join_fc)} to dataset')
                        try:
                            # Copy the specified table to the geodatabase then join to the unioned dataset
                            join_tbl = os.path.join(self.out_gdb, f'{os.path.basename(union_fc)}_jointable')
                            arcpy.management.CopyRows(in_rows=join_fc, out_table=join_tbl)
                            lst_fields = [fld.name for fld in arcpy.ListFields(join_tbl) if fld.name.upper() not in 
                                          ['OBJECTID', 'SHAPE_AREA', 'SHAPE_LENGTH', 'SHAPE']]
                            arcpy.management.JoinField(in_data=union_fc, in_field=lup_ds.source_field, 
                                                       join_table=join_tbl, join_field=lup_ds.join_field, fields=lst_fields)
                        except:
                            self.logger.warning('Something went wrong with the join, could not complete the operation')
                
                
                # Compile the list of fields to search the unioned dataset based on the input schema
                lst_fields = ['SHAPE@AREA', 'SHAPE@LENGTH', fld_id_aoi, fld_id_lu] + \
                                (list(set(lup_ds.assessment_fields) - set(lup_ds.id_fields))) + lup_ds.id_fields
                lst_fields = [fld for fld in lst_fields if fld != 'OBJECTID']
                lst_additional = []
                if lup_ds.other_fields_schema:
                    for fld in lup_ds.other_fields_schema:
                        lst_additional.extend([fld] + lup_ds.other_fields_schema[fld].other_fields)
                    lst_fields.extend(lst_additional)
                if self.fld_aoi:
                    lst_fields.append(self.fld_aoi)

                lst_intersecting_aus = []
                # Gather a list of the assessment units that intersect the area of interest
                with arcpy.da.SearchCursor(in_table=union_fc, field_names=lst_fields, where_clause=lup_ds.sql) as s_cursor:
                    for row in s_cursor:
                        fid_aoi = row[lst_fields.index(fld_id_aoi)]
                        lst_au = []
                        for fld in lup_ds.id_fields:
                            if str(row[lst_fields.index(fld)]) not in ['','None']:
                                lst_au.append(str(row[lst_fields.index(fld)]))
                        if not lst_au:
                            continue
                        au = ' '.join(lst_au)
                        if row[lst_fields.index(fld_id_aoi)] and row[lst_fields.index(fld_id_aoi)] != -1:
                            lst_intersecting_aus.append(au)
                # Flag used to determine if the SQL clause resulted in no records returned
                bl_no_rows = True
                # Loop through the records in the unioned resultant and pull out the required attributes
                with arcpy.da.SearchCursor(in_table=union_fc, field_names=lst_fields, where_clause=lup_ds.sql) as s_cursor:
                    for row in s_cursor:

                        bl_no_rows = False
                        shp = row[lst_fields.index('SHAPE@AREA')]/10000 \
                            if row[lst_fields.index('SHAPE@AREA')] else row[lst_fields.index('SHAPE@LENGTH')]
                        if ds_type == 'Point':
                            shp = 1
                        fid_aoi = row[lst_fields.index(fld_id_aoi)]
                        fid_lup = row[lst_fields.index(fld_id_lu)]
                        aoi = None if not self.fld_aoi else row[lst_fields.index(self.fld_aoi)]
                        lst_au = [] if lup_ds.id_fields else ['All Units']
                        lst_au_name = [] if lup_ds.assessment_fields else ['All Units']

                        # Create the assessment unit id; takes into account if more than one field was selected
                        for fld in lup_ds.id_fields:
                            if str(row[lst_fields.index(fld)]) not in ['', 'None']:
                                lst_au.append(str(row[lst_fields.index(fld)]))
                        if not lst_au:
                            continue
                        else:
                            au = ' '.join(lst_au)

                        # Skip the record if the assessment unit is not in the list of intersecting ones
                        if au not in lst_intersecting_aus and au != 'All Units':
                            continue

                        # Create the assessment unit name; takes into account if more than one field was selected
                        for fld in lup_ds.assessment_fields:
                            if str(row[lst_fields.index(fld)]) not in ['', 'None']:
                                lst_au_name.append(str(row[lst_fields.index(fld)]))
                            else:
                                lst_au_name.append('Unnamed')

                        au_name = ' '.join(lst_au_name)

                        # Incrememnt total area of the assessment unit for overall and the aoi if an aoi field wasselected
                        self.dict_lup_values[lup][ds].aoi[self.str_overall].assessment_units[au].au_name = au_name
                        self.dict_lup_values[lup][ds].aoi[self.str_overall].assessment_units[au].total_area += shp
                        if aoi:
                            self.dict_lup_values[lup][ds].aoi[aoi].assessment_units[au].au_name = au_name
                            self.dict_lup_values[lup][ds].aoi[aoi].assessment_units[au].total_area += shp
                            self.dict_lup_values[lup][ds].aoi[aoi].assessment_units[au].total_count += 1
                        # If the feature is within the area of interest
                        if fid_aoi and fid_aoi != -1 and not pd.isnull(fid_aoi):

                            # Incrememnt total area of the aoi for overall and the aoi if an aoi field was selected
                            self.dict_lup_values[lup][ds].aoi[self.str_overall].total_area += shp
                            self.dict_lup_values[lup][ds].aoi[self.str_overall].total_count += 1
                            if aoi:
                                self.dict_lup_values[lup][ds].aoi[aoi].total_area += shp
                                self.dict_lup_values[lup][ds].aoi[aoi].total_count += 1
                            # If the feature is within the lup value
                            if fid_lup != -1 and not pd.isnull(fid_lup):
                                # Incrememnt area of the aoi within the assessment unit for overall
                                self.dict_lup_values[lup][ds].aoi[self.str_overall].assessment_units[au].aoi_area += shp
                                # Loop through the additional fields and add the values to the dictionary for overall
                                for o_fld in lst_additional:
                                    self.dict_lup_values[lup][ds].aoi[self.str_overall].assessment_units[au].other_fields[o_fld] = row[lst_fields.index(o_fld)] if row[lst_fields.index(o_fld)]else ''
                                # If an aoi field was selected, increment the aoi area within the assessment unit
                                if aoi:
                                    self.dict_lup_values[lup][ds].aoi[aoi].assessment_units[au].aoi_area += shp
                                    # Loop through the additional fields and add the values to the dictionary for theaoi
                                    for o_fld in lst_additional:
                                        self.dict_lup_values[lup][ds].aoi[aoi].assessment_units[au].other_fields[o_fld] = row[lst_fields.index(o_fld)] if row[lst_fields.index(o_fld)] else ''


                # If the flag was not lowered, output warning of no overlap                        
                if bl_no_rows:
                    self.logger.warning(f'+++Value does not overlap the area of interest with the SQL clause: {lup_ds.sql}+++')


    def write_summary(self, ws) -> None:
        i_row = 1
        i_col = 2

        lst_aois = [k for k in self.dict_aoi_area]

        # Add aoi names to the sheet list if there is more than one unique aoi
        # if len(self.dict_aoi_area.keys()) > 1:
        # lst_aois.extend([k for k in self.dict_aoi_area])

        # Write the standardized title information to the top of the sheet
        self.logger.info('Setting up title information')
        title_text = f'Landuse Plan Analysis - {os.path.basename(self.aoi)}'
        ws.cell(row=i_row, column=i_col, value=title_text).style = self.xl_style.title
        ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row, end_column=i_col + 5)
        i_row +=1
        ws.cell(row=i_row, column=i_col, value='File Name/Number:').style = self.xl_style.regular

        ws.cell(row=i_row, column=i_col+1, value=self.file_number).style = self.xl_style.regular
        i_row +=1
        ws.cell(row=i_row, column=i_col, value='Date Submitted:').style = self.xl_style.regular

        ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
        i_row +=1
        ws.cell(row=i_row, column=i_col, value='Submitter Name:').style = self.xl_style.regular

        ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
        i_row +=1
        ws.cell(row=i_row, column=i_col, value='Email:').style = self.xl_style.regular

        ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
        i_row +=1
        ws.cell(row=i_row, column=i_col, value='Ministry/Organization:').style = self.xl_style.regular

        ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
        i_row +=1

        for aoi in lst_aois:
            aoi_value = self.dict_aoi_area[aoi]
            if aoi == self.str_overall:
                aoi_row = i_row
                ws.cell(row=i_row, column=i_col, value='Net AOI Area (ha)*:').style = self.xl_style.regular
            else:
                ws.cell(row=i_row, column=i_col, value=aoi).style = self.xl_style.regular_right

            ws.cell(row=i_row, column=i_col+1, value=aoi_value).style = self.xl_style.number
            i_row += 1

            
        
        ws.cell(row=aoi_row, column=i_col+2, 
                value='*If spatially explicit leave areas are provided they are netted out of the AOI total area (removed). Otherwise, net area = gross area').style = self.xl_style.italics
        ws.merge_cells(start_row=aoi_row, start_column=i_col+2, end_row=aoi_row + len(lst_aois) - 1, end_column=i_col + 7)
        i_row +=2


        lup_headers = ['Category', 'Dataset - Type'] + self.summary_headers
        for header in lup_headers:
            ws.cell(row=i_row, column=i_col + lup_headers.index(header), 
                    value=header).style = self.xl_style.column_header
        
        i_row += 1
        for lup in self.dict_lup_values:
            
            header_text = lup
            i_col = 2
            ws.cell(row=i_row, column=i_col, value=header_text).style = self.xl_style.regular
            row_count = len(self.dict_lup_values[lup]) * len(lst_aois)
            ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row + row_count-1, end_column=i_col)

            
            for ds in self.dict_lup_values[lup]:
                lup_ds = self.dict_lup_values[lup][ds]
                for aoi in lst_aois:
                    if aoi == self.str_overall:
                        ds_value = f'{ds} - {lup_ds.data_type}'
                        ds_style = self.xl_style.regular
                    else:
                        ds_value = aoi
                        ds_style = self.xl_style.regular_aoi
                    i_col = 3
                    i = 1
                    # Gather fields, labels and schema for the specific dataset
                    
                    # Add in a dataset sub header
                    ws.cell(row=i_row, column=i_col, value=ds_value).style = ds_style
                    # if aoi == self.str_overall:
                    #     ws.cell(row=i_row, column=i_col + i, value=lup_ds.data_type).style = self.xl_style.regular
                    #     ws.merge_cells(start_row=i_row, start_column=i_col + i, 
                    #                    end_row=i_row + len(lst_aois) - 1, end_column=i_col + i)
                    # i += 1
                    

                    total_shp = 0
                    aoi_shp = 0
                    au_count = 0
                    for au in lup_ds.aoi[aoi].assessment_units:
                        au_count += 1
                        total_shp += lup_ds.aoi[self.str_overall].assessment_units[au].total_area
                        aoi_shp += lup_ds.aoi[aoi].assessment_units[au].aoi_area
                    if au_count == 0:
                        # If there was no overlap with the aoi, then indicate as such in the sheet and move on to the next  dataset
                        ws.cell(row=i_row, column=i_col + i, 
                                value=f'No overlap with {ds}').style = self.xl_style.regular_na
                        ws.merge_cells(start_row=i_row, start_column=i_col + i, end_row=i_row, 
                                       end_column=i_col + i + 4)
                        

                    else:
                        if aoi == self.str_overall:
                            number_style = self.xl_style.number
                            percent_style = self.xl_style.percent
                        else:
                            number_style = self.xl_style.number_aoi
                            percent_style = self.xl_style.percent_aoi
                        ws.cell(row=i_row, column=i_col + i, value=au_count).style = number_style
                        i += 1
                        ws.cell(row=i_row, column=i_col + i, value=total_shp).style = number_style
                        i += 1
                        ws.cell(row=i_row, column=i_col + i, value=aoi_shp).style = number_style
                        i += 1
                        ws.cell(row=i_row, column=i_col+i, 
                                value=f'=${get_column_letter(i_col+3)}${i_row}/${get_column_letter(i_col+2)}${i_row}'). style=percent_style
                        i += 1

                        if lup_ds.data_type not in ['Point', 'Polyline']:
                            value = f'=${get_column_letter(i_col+3)}${i_row}/${get_column_letter(3)}${aoi_row + lst_aois.index(aoi)}'
                            style = percent_style
                        else:
                            value = 'N/A'
                            style = self.xl_style.regular_na

                        ws.cell(row=i_row, column=i_col+i, value=value).style=style
                    i_row += 1

                

        
        i_col = 1
        # Set the column widths
        ws.column_dimensions[get_column_letter(i_col)].width = 1
        ws.column_dimensions[get_column_letter(i_col+1)].width = 40
        ws.column_dimensions[get_column_letter(i_col+2)].width = 40
        ws.column_dimensions[get_column_letter(i_col+3)].width = 20
        ws.column_dimensions[get_column_letter(i_col+4)].width = 40
        ws.column_dimensions[get_column_letter(i_col+5)].width = 40
        ws.column_dimensions[get_column_letter(i_col+6)].width = 30
        ws.column_dimensions[get_column_letter(i_col+7)].width = 30
        ws.column_dimensions[get_column_letter(i_col+8)].width = 30


            


    def write_excel(self) -> None:
        """
        CLASS METHOD

        write_excel: Function that is used to produce the excel analysis report based on the information found from the overlays, and using the formatting options laid out in the schems document
        """

        # Check if the output excel file already exists, delete if it does
        self.logger.info('Writing excel output')
        if os.path.exists(self.xls_output):
            os.remove(self.xls_output)

        # Copy the schema excel to the new location and remove all sheets but the front page
        # shutil.copyfile(self.xls_schema, self.xls_output)
        # wb = openpyxl.load_workbook(self.xls_output)
        # for sheet in wb.sheetnames:
        #     if sheet != 'Front Page':
        #         wb.remove(worksheet=wb[sheet])
        wb = openpyxl.Workbook()
        for sheet in wb.sheetnames:
            if sheet == 'Sheet':
                wb.remove(worksheet=wb[sheet])

        # Set up the standard styles and add to the workbook
        self.xl_style = ExcelStyles(wb=wb)
        lst_sheets = [self.str_summary]

        # Add aoi names to the sheet list if there is more than one unique aoi
        if len(self.dict_aoi_area.keys()) > 1:
            lst_sheets.extend([k for k in self.dict_aoi_area])

        try:
        # Loop through the list of sheets to create
            for sheet in lst_sheets:

                # Add a new sheet with the specified name and set the worksheet opject to point to it
                self.logger.info(f'Writing sheet - {sheet}')
                wb.create_sheet(title=str(sheet))
                ws = wb[str(sheet)]

                if sheet == self.str_summary:
                    self.write_summary(ws=ws)
                    continue

                # Row and column index objects
                i_row = 1
                i_col = 2

                # Write the standardized title information to the top of the sheet
                self.logger.info('Setting up title information')
                title_text = f'Landuse Plan Analysis - {os.path.basename(self.aoi)}'
                ws.cell(row=i_row, column=i_col, value=title_text).style = self.xl_style.title
                ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row, end_column=i_col + 5)
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='File Name/Number:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value=self.file_number).style = self.xl_style.regular
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='Date Submitted:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='Submitter Name:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='Email:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='Ministry/Organization:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular
                i_row +=1
                ws.cell(row=i_row, column=i_col, value='Net AOI Area (ha)*:').style = self.xl_style.regular

                ws.cell(row=i_row, column=i_col+1, value='').style = self.xl_style.regular

                # Use the overall area if the sheet is overall, otherwise use the area value for the specific aoi part
                aoi_value = self.dict_aoi_area[sheet]
                ws.cell(row=i_row, column=i_col+1, value=aoi_value).style = self.xl_style.number
                aoi_row = i_row
                ws.cell(row=i_row, column=i_col+2, 
                        value='*If spatially explicit leave areas are provided they are netted out of the AOI total area (removed). Otherwise, net area = gross area').style = self.xl_style.italics
                ws.merge_cells(start_row=i_row, start_column=i_col+2, end_row=i_row, end_column=i_col + 6)

                i_row +=2

                # Loop through the lup values in the dictionary
                for lup in self.dict_lup_values:
                    i_col = 2
                    start_col = i_col
                    start_row = i_row
                    self.logger.info(f'Writing {lup} results')
                    # Write the header text
                    header_text = lup
                    num_fields = 0
                    ws.cell(row=i_row, column=i_col, value=header_text).style = self.xl_style.value_header

                    for ds in self.dict_lup_values[lup]:
                        num_fields = len(self.dict_lup_values[lup][ds].other_fields_schema) if len(self.dict_lup_values[lup][ds].other_fields_schema) > num_fields else num_fields
                    ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row, 
                                       end_column=i_col + len(self.standard_poly_headers) + num_fields + 1)
                    end_col = i_col + len(self.standard_poly_headers) + num_fields + 1
                    i_row += 1
                    
                    # Loop through the datasets for the specified lup value
                    for ds in self.dict_lup_values[lup]:
                        i_col = 3
                        # Gather fields, labels and schema for the specific dataset
                        lup_ds = self.dict_lup_values[lup][ds]
                        ds_merge_count = 1 if len(lup_ds.aoi[sheet].assessment_units) == 0 \
                            else len(lup_ds.aoi[sheet].assessment_units)
                        # Add in a dataset sub header
                        ws.cell(row=i_row, column=i_col, value=ds).style = self.xl_style.value_subheader
                        ws.merge_cells(start_row=i_row, start_column=i_col, 
                                       end_row=i_row + ds_merge_count, end_column=i_col)
                        i_col += 1
                        
                        lup_fields = list(lup_ds.other_fields_schema.keys())
                        other_headers = [lup_ds.other_fields_schema[fld].label for fld in lup_ds.other_fields_schema]

                        standard_headers = self.standard_point_headers if lup_ds.data_type == 'Point' else \
                            self.standard_line_headers if lup_ds.data_type == 'Polyline' else self.standard_poly_headers
                        lup_headers = [standard_headers[0]] + other_headers + standard_headers[1:]
                        add_index = i_col + len(lup_fields)

                        column_length = len(lup_headers)

                        # Write the standardized column headers and any additional ones specified in the schema
                        for header in lup_headers:
                            ws.cell(row=i_row, column=i_col + lup_headers.index(header), 
                                    value=header).style = self.xl_style.column_header
                        i_row += 1

                        
                        au_count = 0
                        # Loop through the assessment units for the dataset
                        for au in lup_ds.aoi[sheet].assessment_units:
                            au_count += 1
                            assess_unit = lup_ds.aoi[sheet].assessment_units[au]
                            i = 1
                            # Write the standardized values in the first 5 columns
                            ws.cell(row=i_row, column=i_col, value=assess_unit.au_name).style=self.xl_style.regular
                            ws.cell(row=i_row, column=add_index+i, 
                                    value=lup_ds.aoi[self.str_overall].assessment_units[au].total_area).style=self.xl_style.number
                            i += 1

                            ws.cell(row=i_row, column=add_index+i, 
                                    value=assess_unit.aoi_area).style=self.xl_style.number
                            i += 1

                            ws.cell(row=i_row, column=add_index+i, 
                                    value=f'=${get_column_letter(add_index+2)}${i_row}/${get_column_letter(add_index+1)}${i_row}').style=self.xl_style.percent
                            i += 1

                            if lup_ds.data_type not in ['Point', 'Polyline']:
                                value = f'=${get_column_letter(add_index+2)}${i_row}/${get_column_letter(3)}${aoi_row}'
                                style = self.xl_style.percent
                            else:
                                value = 'N/A'
                                style = self.xl_style.regular_na
                            
                            ws.cell(row=i_row, column=add_index+i, value=value).style=style

                            # Loop through the additional fields for the dataset
                            for ce_fld in lup_fields:
                                # Pull the values from the schema object
                                ce_schema = lup_ds.other_fields_schema[ce_fld]
                                col_index = i_col + lup_headers.index(ce_schema.label)
                                try:
                                    ce_value = lup_ds.aoi[sheet].assessment_units[au].other_fields[ce_fld]
                                except:
                                    continue

                                # Convert to a float rounded to two decimal places if the value is a decimal
                                try:
                                    flt_value = float(ce_value)
                                    ce_value = round(ce_value, 3) if '.' in str(ce_value) else ce_value
                                except:
                                    pass
                                ws.cell(row=i_row, column=col_index, value=ce_value)

                                # If there isn't a value, then keep the cell style as regular
                                if not ce_value:
                                    ws.cell(row=i_row, column=col_index).style = self.xl_style.regular
                                    continue

                                # Gather the other field information and add it within brackets after the main value
                                if ce_schema.other_fields:
                                    lst_other_values = [lup_ds.aoi[sheet].assessment_units[au].other_fields[o_fld] for o_fld in ce_schema.other_fields]
                                    join_val = f'{ws.cell(row=i_row, column=col_index).value} ({",".join(lst_other_values)})'
                                    ws.cell(row=i_row, column=col_index, value=join_val)

                                # Pull the formatting values from the schema object and create the cell style
                                cell_style = self.xl_style.regular
                                if ce_schema.value_type:
                                    # If the value type is discrete, pull the values as is based on the data
                                    if ce_schema.value_type == 'Discrete':
                                        val_style = ce_schema.dict_values[ce_value]

                                    # If the value type is a range, then cycle through the different range types and assign     the format if it meets the requirements
                                    elif ce_schema.value_type == 'Range':
                                        for val_key in ce_schema.dict_values:
                                            val_schema = ce_schema.dict_values[val_key]
                                            if val_schema.range_low and val_schema.range_high:
                                                if val_schema.range_low <= ce_value <= val_schema.range_high:
                                                    break
                                            elif not val_schema.range_low and val_schema.range_high:
                                                if val_schema.range_operator == '<=' and ce_value <= val_schema.range_high:
                                                    break
                                                elif val_schema.range_operator == '<' and ce_value < val_schema.range_high:
                                                    break
                                            elif val_schema.range_low and not val_schema.range_high:
                                                if val_schema.range_operator == '>=' and ce_value >= val_schema.range_low:
                                                    break
                                                elif val_schema.range_operator == '>' and ce_value > val_schema.range_low:
                                                    break
                                        val_style = val_schema


                                    # Create a new style object based on the extracted formats, then assign it to the cell
                                    cell_style = self.xl_style.create_style_copy(wb=wb, 
                                                                                 name=f'{lup}-{ds}-{sheet}-{ce_fld}-    {ce_value}', font=val_style.style_font, 
                                                                                 align=val_style.style_align,   border=val_style.style_border,    fill=val_style.style_fill,     num_format=val_style.style_format)

                                ws.cell(row=i_row, column=col_index).style = cell_style
                            end_row = i_row
                            i_row += 1

                        # If there was no overlap with the aoi, then indicate as such in the sheet and mvoe on to the next  dataset
                        if au_count == 0:
                            ws.cell(row=i_row, column=i_col, value=f'No overlap with {ds}').style = self.xl_style.regular_na
                            ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row, 
                                       end_column=i_col + column_length-1)
                            end_row = i_row
                            i_row += 1
                        
                        i_row += 1
                    
                    # If there was no overlap with the aoi, then indicate as such in the sheet and mvoe on to the next  dataset
                    # if len(self.dict_lup_values[lup][ds].aoi[sheet].assessment_units) == 0:
                    #     i_row -= 1
                    #     ws.cell(row=i_row, column=i_col, value=f'No overlap with {ds}').style = self.xl_style.regular
                    #     ws.merge_cells(start_row=i_row, start_column=i_col, end_row=i_row, 
                    #                    end_column=i_col + column_length-1)
                    #     end_row = i_row
                    #     i_row += 3
                        

                    for row_idx in range(start_row, end_row + 1):
                        for col_idx in range(start_col, end_col + 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cur_border = copy(cell.border) if cell.border else Border()

                            if row_idx == start_row:
                                cur_border.top = Side(style='thick')
                            if row_idx == end_row:
                                cur_border.bottom = Side(style='thick')
                            if col_idx == start_col:
                                cell.style = self.xl_style.value_header
                                cur_border.left = Side(style='thick')
                            if col_idx == end_col:
                                cur_border.right = Side(style='thick')
                            if col_idx == start_col and row_idx == start_row:
                                cur_border.bottom = None
                            cell.border = cur_border


                i_col = 1
                # Set the column widths
                ws.column_dimensions[get_column_letter(i_col)].width = 1
                ws.column_dimensions[get_column_letter(i_col+1)].width = 20
                ws.column_dimensions[get_column_letter(i_col+2)].width = 60
                ws.column_dimensions[get_column_letter(i_col+3)].width = 45
                for i in range(i_col+4, i_col+12):
                    ws.column_dimensions[get_column_letter(i)].width = 25

        except Exception as e:
            self.logger.error(f'Error creating the excel: {e}')
            self.logger.error(traceback.format_exc())
            wb.save(self.xls_output)
            sys.exit()
        # Activate the first sheet and save the output
        wb.active = wb.worksheets[0]
        wb.save(self.xls_output)


class LU_Value:
    """
    CLASS

    Support class object that contains all information about LU Values as pulled from the schema
    """
    def __init__(self, name:str, category:str, id_fields: list=[], assess_fields:list=[], 
                 path:str='', sql:str=None, source_field:str=None, join_table_path:str=None, 
                 join_table_type:str=None, join_table_field:str=None, buffer: float=None) -> None:
        """
        CLASS METHOD
        
        __init__: Initializes that class object and sets up the class properties

        Args:
            name (str): Name of the lup value dataset
            category (str): Category the dataset belongs to
            unique_field (str, optional): Unique ID field. Defaults to None.
            assess_field (str, optional): Assessment unit field. Defaults to None.
            path (str, optional): File path of the dataset. Defaults to ''.
            sql (str, optional): Definition query of the dataset. Defaults to None.
            source_field (str, optional): Source field if joining a table to this dataset. Defaults to None.
            join_table_path (str, optional): Join table path if joining to the dataset. Defaults to None.
            join_table_field (str, optional): Join field within the join table if joining. Defaults to None.
            buffer (float, optional): value to buffer the input dataset by. Defaults to None
        """
        
        self.name = name
        self.category = category
        self.id_fields = id_fields
        self.assessment_fields = assess_fields
        self.no_id = False
        if not self.id_fields and self.assessment_fields:
            self.id_fields = self.assessment_fields
        elif not self.assessment_fields and self.id_fields:
            self.assessment_fields = self.id_fields
        elif not self.assessment_fields and not self.id_fields:
            self.no_id = True
        self.path = path
        self.data_type = None
        self.sql = sql
        self.source_field = source_field
        self.join_path = join_table_path
        self.join_field = join_table_field
        self.join_table_type = join_table_type
        self.buffer = buffer
        self.fc_union = None
        self.aoi = defaultdict(AOI) # Dictionary of AOI objects
        self.other_fields_schema = defaultdict(FieldSchema) # Dictionary of FieldSchema objects

class AOI:
    """ 
    CLASS
    
    Support class to contain information applicable to area of interests
    """
    def __init__(self) -> None:
        """
        CLASS METHOD

        __init__: Initializes the properties associated with the class object
        """
        self.total_area = 0
        self.total_count = 0
        self.assessment_units = defaultdict(Assessment_Unit) # Dictionary of Assessment_Unit objects
    


class Assessment_Unit:
    """
    CLASS

    Support class to contain information applicable to assessment units
    """
    def __init__(self) -> None:
        """
        CLASS METHOD

        __init__: Initializes the properties associated with the class object
        """
        self.total_area = 0
        self.total_count = 0
        self.aoi_area = 0
        self.au_name = ''
        self.other_fields = defaultdict()

class FieldSchema:
    """
    CLASS

    Support class to contain information applicable to field schemas
    """
    def __init__(self) -> None:
        """
        CLASS METHOD

        __init__: Initializes the properties associated with the class object
        """
        self.name = ''
        self.label = ''
        self.dict_values = defaultdict(ValueSchema) # Dictionary of ValueSchema objects
        self.value_type = ''
        self.other_fields = []

class ValueSchema:
    """
    CLASS

    Support class to contain information applicable to value schemas
    """
    def __init__(self) -> None:
        """
        CLASS METHOD

        __init__: Initializes the properties associated with the class object
        """
        self.discrete_value = ''
        self.range_high = 0
        self.range_low = 0
        self.range_operator = None
        self.label = ''
        self.style_font = Font # OpenPyXl font object
        self.style_align = Alignment # OpenPyXl alignment object
        self.style_border = Border # OpenPyXl border object
        self.style_fill = PatternFill # OpenPyXl patternfill object
        self.style_format = ''

class ExcelStyles:
    """
    CLASS

    Support class to contain information and methods applicable to OpenPyXl workbook styes
    """
    def __init__(self, wb: openpyxl.Workbook) -> None:
        """
        CLASS METHOD

        __init__: Initializes the properties associated with the class object

        Args:
            wb (openpyxl.Workbook): OpenPyXl workbook object
        """
        self.thin_border = Side(style='thin', color='000000')
        self.wb = wb

        # Create the standard styles in the workbook upon creation of the class object
        self.title = self.create_style(wb=self.wb, name='title', bold=True, font_size=12, horiz_align='left')
        self.value_header = self.create_style(wb=self.wb, name='value header', bold=True, font_size=11,
                                              cell_border=True, cell_fill='e2e7a7', horiz_align='left')
        self.value_subheader = self.create_style(wb=self.wb, name='value subheader', bold=True, italic=True,
                                                 cell_border=True, cell_fill='f9faed', vert_align='top', horiz_align='left')
        self.regular = self.create_style(wb=self.wb, name='regular', cell_border=True, font_size=8, horiz_align='left')
        self.regular_na = self.create_style(wb=self.wb, name='regular na', cell_border=True, font_size=8, 
                                            horiz_align='left', cell_fill='808080', italic=True)
        self.regular_right = self.create_style(wb=self.wb, name='regular right', cell_border=True, horiz_align='right', font_size=8)
        self.italics = self.create_style(wb=self.wb, name='italics', font_size=8, italic=True, horiz_align='left', 
                                         cell_border=True)
        self.percent = self.create_style(wb=self.wb, name='regular percent', cell_border=True, cell_format='##0.00%', font_size=8, horiz_align='right')
        self.number = self.create_style(wb=self.wb, name='regular number', horiz_align='right', cell_border=True, 
                                        cell_format='###,##0', font_size=8)
        self.decimal = self.create_style(wb=self.wb, name='decimal number', horiz_align='right', cell_border=True, 
                                         cell_format='###,##0.0', font_size=8)
        self.column_header = self.create_style(wb=self.wb, name='column header', bold=True, cell_border=True, cell_fill='edfcfc')
        self.regular_aoi = self.create_style(wb=self.wb, name='regular aoi', cell_border=True, horiz_align='right',
                                              font_size=8, italic=True, cell_fill='e8e8e8')
        self.number_aoi = self.create_style(wb=self.wb, name='number aoi', horiz_align='right', cell_border=True, 
                                        cell_format='###,##0', font_size=8, italic=True, cell_fill='e8e8e8')
        self.percent_aoi = self.create_style(wb=self.wb, name='percent aoi', cell_border=True, cell_format='##0.00%', 
                                             font_size=8, horiz_align='right', italic=True, cell_fill='e8e8e8')


    def create_style(self, wb:openpyxl.Workbook, name: str, font_size: int=10, bold=False, italic=False, 
                     text_colour: str='000000', horiz_align: str='center', vert_align: str='center', 
                     wrap_text: bool=True, cell_border: bool=False, cell_fill: str=None, 
                     cell_format=None) -> NamedStyle:
        """
        CLASS METHOD

        create_style: Function that creates a style object based on the parameters passed in

        Args:
            wb (openpyxl.Workbook): OpenPyXl workbook object
            name (str): name of the style
            font_size (int, optional): font size of the cell text. Defaults to 10.
            bold (bool, optional): Is the cell bold. Defaults to False.
            italic (bool, optional): Is the cell italic. Defaults to False.
            text_colour (str, optional): Colour of the text. Defaults to '000000'.
            horiz_align (str, optional): horizontal alignment of the cell. Defaults to 'center'.
            vert_align (str, optional): vertical alignment of the cell. Defaults to 'center'.
            wrap_text (bool, optional): Is the cell text wrapped. Defaults to True.
            cell_border (bool, optional): Is there a border on the cell. Defaults to False.
            cell_fill (str, optional): Background colour of the cell. Defaults to None.
            cell_format (_type_, optional): Number format of the cell. Defaults to None.

        Returns:
            NamedStyle: OpenPyXl Named style object created from the input parameters
        """
        
        # Create the new style
        new_style = NamedStyle(name=name)
        new_style.font = Font(size=font_size, bold=bold, italic=italic, color=text_colour, name='Calibri')
        new_style.alignment = Alignment(horizontal=horiz_align, vertical=vert_align, wrap_text=wrap_text)
        if cell_border:
            new_style.border = Border(left=self.thin_border, top=self.thin_border, right=self.thin_border, 
                                      bottom=self.thin_border)
        if cell_fill:
            new_style.fill = PatternFill(patternType='solid', start_color=cell_fill)
        if cell_format:
            new_style.number_format = cell_format

        # Add the style to the workbook
        wb.add_named_style(style=new_style)

        return new_style
    
    def create_style_copy(self, wb:openpyxl.Workbook, name: str, font: Font, align: Alignment, border: Border, 
                          fill: PatternFill, num_format: str) -> NamedStyle:
        """
        CLASS METHOD

        create_style_copy: Creates a style object based on OpenPyXl style objects passed in

        Args:
            wb (openpyxl.Workbook): OpenPyXl workbook object
            name (str): name of the style
            font (Font): OpenPyXl font object
            align (Alignment): OpenPyXl alignment object
            border (Border): OpenPyXl border object
            fill (PatternFill): OpenPyXl PatternFill object
            num_format (str): Number format

        Returns:
            NamedStyle: OpenPyXl Named style object created from the input parameters
        """

        # Delete the named style if it already exists in the workbook
        if name in wb.named_styles:
            del wb._named_styles[wb.style_names.index(name)]

        # Create new style object and assign properties

        new_style = NamedStyle(name=name)
        new_style.font = font
        new_style.alignment = align
        new_style.alignment.wrap_text = True
        new_style.border = border
        new_style.fill = fill
        new_style.number_format = num_format

        # Add the style to the workbook
        wb.add_named_style(style=new_style)

        return new_style


# First line the script hits when run, passes control up to the run_app function
if __name__ == '__main__':
    run_app()
